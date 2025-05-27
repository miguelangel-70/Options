import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import os
import io
import openpyxl
import shutil

# ===================== CONFIGURACIÓN DE LA APLICACIÓN =====================
st.set_page_config(
    page_title="Dashboard de Volatilidad",
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_DATOS_PATH = "base_datos.xlsx"

# ==== Funciones auxiliares para hojas ====
def obtener_nombre_hoja(path_excel):
    try:
        wb = openpyxl.load_workbook(path_excel, read_only=True)
        return wb.sheetnames[0]
    except Exception:
        return "Datos"

def obtener_hojas(path_excel):
    try:
        wb = openpyxl.load_workbook(path_excel, read_only=True)
        return wb.sheetnames
    except Exception:
        return []

# ==== Estado inicial ====
if "nombre_base_activa" not in st.session_state:
    st.session_state["nombre_base_activa"] = None
if "df_memoria" not in st.session_state:
    st.session_state["df_memoria"] = None
if "nombre_hoja_excel" not in st.session_state:
    st.session_state["nombre_hoja_excel"] = "Datos"
if "hojas_disponibles" not in st.session_state:
    st.session_state["hojas_disponibles"] = []

# ==== Carga automática de base de datos ====
if st.session_state["df_memoria"] is None and os.path.exists(BASE_DATOS_PATH):
    try:
        hojas_disponibles = obtener_hojas(BASE_DATOS_PATH)
        if hojas_disponibles:
            hoja_excel = hojas_disponibles[0]
            df_base = pd.read_excel(BASE_DATOS_PATH, sheet_name=hoja_excel)
            if 'Call Open Interest' not in df_base.columns or 'Put Open Interest' not in df_base.columns:
                df_base.rename(columns={
                    'Open Interest': 'Call Open Interest',
                    'Open Interest.1': 'Put Open Interest'
                }, inplace=True)
            df_base['Expiration Date'] = pd.to_datetime(df_base['Expiration Date'])
            df_base['Fecha de Extracción'] = pd.to_datetime(df_base['Fecha de Extracción']).dt.date
            st.session_state["df_memoria"] = df_base
            st.session_state["nombre_base_activa"] = BASE_DATOS_PATH
            st.session_state["hojas_disponibles"] = hojas_disponibles
            st.session_state["nombre_hoja_excel"] = hoja_excel
    except Exception:
        st.session_state["df_memoria"] = None
        st.session_state["nombre_base_activa"] = None

# ===================== CARGAR CSS =====================
if os.path.exists("style.css"):
    with open("style.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ===================== MENÚ PRINCIPAL =====================
with st.sidebar:
    if st.session_state.get("hojas_disponibles"):
        st.markdown("<h4 style='color: white; margin-bottom: 0.5rem;'>📄 Selección de hoja activa</h4>",unsafe_allow_html=True)

        hoja_seleccionada = st.selectbox(
            "Selecciona hoja activa",  # Etiqueta accesible
            st.session_state["hojas_disponibles"],
            index=st.session_state["hojas_disponibles"].index(
                st.session_state.get("nombre_hoja_excel", "Datos")
            ),
            key="hoja_activa",
            label_visibility="collapsed"  # Oculta visualmente la etiqueta, pero mantiene accesibilidad
        )
        # Detectar cambio de hoja y actualizar df_memoria si es necesario

    # Detectar cambio de hoja y actualizar df_memoria si es necesario
    if hoja_seleccionada != st.session_state.get("nombre_hoja_excel"):
        st.session_state["nombre_hoja_excel"] = hoja_seleccionada
        try:
            df = pd.read_excel(BASE_DATOS_PATH, sheet_name=hoja_seleccionada)
            st.session_state["df_memoria"] = df
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo cargar la hoja seleccionada: {str(e)}")
            
    selected = option_menu(
        "Menú Principal",
        ["Visualización", "Cargar Datos", "Configuración"],
        icons=["bar-chart", "upload", "gear"],
        menu_icon="cast",
        default_index=0,
    )

# ===================== FUNCIONES =====================
def mostrar_mensaje(tipo, texto):
    if tipo == "success":
        st.toast(texto, icon="✅")
    elif tipo == "warning":
        st.toast(texto, icon="⚠️")
    elif tipo == "error":
        st.toast(texto, icon="❌")
    else:
        st.toast(texto, icon="ℹ️")

def cargar_xlsx(file):
    try:
        # Crear backup si existe una base de datos previa
        if os.path.exists(BASE_DATOS_PATH):
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"base_datos_backup_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copyfile(BASE_DATOS_PATH, backup_path)
            mostrar_mensaje("success", f"Copia de seguridad creada: {backup_path}")

        # Obtener las hojas del archivo cargado
        hojas = obtener_hojas(file)
        if not hojas:
            mostrar_mensaje("error", "No se encontraron hojas en el archivo.")
            return None

        hoja = hojas[0]
        df = pd.read_excel(file, sheet_name=hoja)

        # Renombrar columnas si es necesario
        if 'Call Open Interest' not in df.columns or 'Put Open Interest' not in df.columns:
            df.rename(columns={
                'Open Interest': 'Call Open Interest',
                'Open Interest.1': 'Put Open Interest'
            }, inplace=True)

        # Convertir fechas
        df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        df['Fecha de Extracción'] = pd.to_datetime(df['Fecha de Extracción']).dt.date

        # Guardar el archivo cargado como nueva base de datos
        with open(BASE_DATOS_PATH, "wb") as out_file:
            out_file.write(file.getbuffer())

        # Actualizar hojas disponibles (leer desde el nuevo archivo guardado)
        hojas_actualizadas = obtener_hojas(BASE_DATOS_PATH)

        # Actualizar estado de sesión
        st.session_state["df_memoria"] = df
        st.session_state["nombre_base_activa"] = file.name if hasattr(file, 'name') else BASE_DATOS_PATH
        st.session_state["nombre_hoja_excel"] = hoja
        st.session_state["hojas_disponibles"] = hojas_actualizadas

        mostrar_mensaje("success", f"Archivo Excel cargado correctamente y guardado como base de datos.")

        return df

    except Exception as e:
        mostrar_mensaje("error", f"Error al cargar y guardar el archivo Excel: {str(e)}")
        return None

def cargar_csv(file):
    try:
        df = pd.read_csv(file, skiprows=3, header=0, sep=',')
        df = df[['Expiration Date', 'Open Interest', 'Strike', 'Open Interest.1']]
        df.columns = ['Expiration Date', 'Call Open Interest', 'Strike', 'Put Open Interest']
        df.insert(0, 'Fecha de Extracción', pd.to_datetime('today').date())
        #df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        return df
    except Exception as e:
        mostrar_mensaje("error", f"Error al cargar el archivo CSV: {str(e)}")
        return None
    
def guardar_base_datos(df, hoja_destino, backup=False, backup_name=None):
    try:
        # Leer todas las hojas existentes, excepto la que será reemplazada
        hojas = {}
        if os.path.exists(BASE_DATOS_PATH):
            try:
                xls = pd.ExcelFile(BASE_DATOS_PATH, engine="openpyxl")
                for hoja in xls.sheet_names:
                    if hoja != hoja_destino:
                        hojas[hoja] = xls.parse(hoja)
            except Exception as e:
                mostrar_mensaje("error", f"No se pudieron leer las hojas existentes: {e}")
                hojas = {}

        # Guardar la hoja destino con sus nuevos datos
        hojas[hoja_destino] = df

        with pd.ExcelWriter(BASE_DATOS_PATH, engine="openpyxl", mode="w") as writer:
            for hoja, datos in hojas.items():
                datos.to_excel(writer, sheet_name=hoja, index=False)

        # Crear backup si se solicita
        if backup and os.path.exists(BASE_DATOS_PATH):
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = backup_name or f"base_datos_backup_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copyfile(BASE_DATOS_PATH, backup_path)
            mostrar_mensaje("success", f"Copia de seguridad completa creada: {backup_path}")

        # Leer nuevamente las hojas desde el archivo para evitar duplicados
        st.session_state["hojas_disponibles"] = obtener_hojas(BASE_DATOS_PATH)
        mostrar_mensaje("success", f"Datos guardados exitosamente en la hoja '{hoja_destino}'")

    except Exception as e:
        mostrar_mensaje("error", f"Error al guardar la base de datos: {str(e)}")

def generar_grafico_barras(df, fecha_vencimiento, fecha_carga=None):
    """Genera un gráfico de barras horizontales para los TOP 10 CALL y TOP 10 PUT con etiquetas de valor."""

    top_calls = df.nlargest(10, 'Call Open Interest')
    top_puts = df.nlargest(10, 'Put Open Interest')

    strikes_top = pd.concat([top_calls, top_puts]).drop_duplicates(subset='Strike').sort_values('Strike')

    max_call = strikes_top['Call Open Interest'].max()
    max_put = strikes_top['Put Open Interest'].max()
    max_oi = max(max_call, max_put)
    x_buffer = max_oi * 0.2

    fig, ax = plt.subplots(figsize=(12, 7))

    if fecha_carga:
        plt.suptitle(
            f"{hoja_seleccionada} - Open Interest (TOP 10 CALL & TOP 10 PUT)\n"
            f"Vencimiento: {fecha_vencimiento.strftime('%Y-%m-%d')} | "
            f"Extracción: {fecha_carga.strftime('%Y-%m-%d')}",
            fontsize=15
        )
    else:
        plt.suptitle(
            f"{hoja_seleccionada} - Open Interest (TOP 10 CALL & TOP 10 PUT)\n"
            f"Vencimiento: {fecha_vencimiento.strftime('%Y-%m-%d')}",
            fontsize=15
        )

    plt.title(f"Informe generado el: {datetime.now().strftime('%Y-%m-%d')}", fontsize=11, style='italic', loc='center')

    # Dibujar barras
    bar_labels = strikes_top['Strike'].astype(str)
    call_values = -strikes_top['Call Open Interest']
    put_values = strikes_top['Put Open Interest']

    bars_call = ax.barh(bar_labels, call_values, color='#4CAF50', alpha=0.8, label='CALL OI')
    bars_put = ax.barh(bar_labels, put_values, color='#F44336', alpha=0.8, label='PUT OI')

    ax.axvline(0, color='black', linestyle='-', linewidth=1.5)

    # Añadir valores CALL
    for bar in bars_call:
        width = bar.get_width()
        if width != 0:
            ax.annotate(
                f'{abs(int(width)):,}',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(-5, 0),
                textcoords="offset points",
                ha='right',
                va='center',
                color='#2E7D32',
                fontsize=10
            )

    # Añadir valores PUT
    for bar in bars_put:
        width = bar.get_width()
        if width != 0:
            ax.annotate(
                f'{int(width):,}',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(5, 0),
                textcoords="offset points",
                ha='left',
                va='center',
                color='#C62828',
                fontsize=10
            )

    ax.set_xlim(-max_call - x_buffer, max_put + x_buffer)
    ax.set_xlabel('Open Interest', labelpad=10)
    ax.set_ylabel('Strike Price', labelpad=10)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), ncol=2, framealpha=1)
    ax.grid(axis='y', linestyle='--', alpha=0.4)

    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)

    fig.tight_layout()
    return fig

# ===================== FUNCIONALIDAD =====================
if selected == "Cargar Datos":
    st.markdown("<h2 class='fade-in'>Cargar Datos</h2>", unsafe_allow_html=True)
    
    # Mostrar mensaje de permisos y detener la ejecución
    st.error("🔒 Necesarios permisos de administrador para modificar la base de datos")
    st.info("Contacte con el administrador del sistema para realizar modificaciones en la base de datos.")
    st.stop()

elif selected == "Visualización":
    st.markdown("<h2 class='fade-in'>Visualización</h2>", unsafe_allow_html=True)
    df = st.session_state["df_memoria"]

    if df is not None and not df.empty:
        # Filtrar fechas de extracción que tengan al menos una fecha de vencimiento asociada
        fechas_extraccion_validas = df.groupby('Fecha de Extracción')['Expiration Date'].nunique()
        fechas_extraccion = sorted(fechas_extraccion_validas[fechas_extraccion_validas > 0].index, reverse=True)

        # Inicialmente se puede dejar las fechas de vencimiento vacías (se actualizarán luego)
        fechas_vencimiento = []


        col1, col2 = st.columns(2)

        with col1:
            fecha_extraccion = st.selectbox(
                "Seleccione fecha de extracción:",
                fechas_extraccion,
                format_func=lambda x: x.strftime('%Y-%m-%d')
            )

        with col2:
            fechas_vencimiento = sorted(df[df['Fecha de Extracción'] == fecha_extraccion]['Expiration Date'].unique())
            fecha_vencimiento = st.selectbox(
                "Seleccione fecha de vencimiento:",
                fechas_vencimiento,
                format_func=lambda x: x.strftime('%Y-%m-%d')
            )

        df_filtrado = df[(df['Fecha de Extracción'] == fecha_extraccion) & (df['Expiration Date'] == fecha_vencimiento)]

        if not df_filtrado.empty:
            st.subheader(f"{st.session_state['nombre_hoja_excel']} - Open Interest\nExtracción: {fecha_extraccion} | Vencimiento: {fecha_vencimiento}")
            fig = generar_grafico_barras(df_filtrado, fecha_vencimiento, fecha_extraccion)
            st.pyplot(fig)

            buf = io.BytesIO()
            fig.savefig(buf, format="png", bbox_inches='tight')
            buf.seek(0)

            file_name = (
                f"{st.session_state['nombre_hoja_excel']} - IO vencimiento ({fecha_vencimiento.strftime('%Y-%m-%d')}) "
                f"- extraccion ({fecha_extraccion.strftime('%Y-%m-%d')}).png"
            )
            
            st.download_button("📥 Descargar imagen", data=buf, file_name=file_name, mime="image/png")

            if st.checkbox("Mostrar tabla de datos completos", value=False):
                st.dataframe(df_filtrado[['Strike', 'Call Open Interest', 'Put Open Interest']].sort_values('Strike'))
        else:
            mostrar_mensaje("warning", "No hay datos disponibles para los filtros seleccionados.")
    else:
        st.info("No se ha cargado ninguna base de datos. Por favor, use la opción 'Cargar Datos'.")

elif selected == "Configuración":
    st.markdown("<h2 class='fade-in'>Configuración</h2>", unsafe_allow_html=True)
    st.write("Opciones de configuración próximamente.")
