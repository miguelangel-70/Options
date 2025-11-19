import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import os
import io
import openpyxl
import shutil
import sys

# ===================== CONFIGURACIÓN DE LA APLICACIÓN =====================
st.set_page_config(
    page_title="Dashboard de Volatilidad",
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_DATOS_PATH = "base_datos.xlsx"

# ==== Funciones auxiliares para hojas ====
def obtener_nombre_hoja(path_excel):
    try:
        wb = openpyxl.load_workbook(path_excel, read_only=True)
        return wb.sheetnames[0]
    except Exception:
        return "Datos"

def obtener_hojas(path_excel):
    try:
        wb = openpyxl.load_workbook(path_excel, read_only=True)
        return wb.sheetnames
    except Exception:
        return []

# ==== Estado inicial ====
if "nombre_base_activa" not in st.session_state:
    st.session_state["nombre_base_activa"] = None
if "df_memoria" not in st.session_state:
    st.session_state["df_memoria"] = None
if "nombre_hoja_excel" not in st.session_state:
    st.session_state["nombre_hoja_excel"] = "Datos"
if "hojas_disponibles" not in st.session_state:
    st.session_state["hojas_disponibles"] = []
if "confirmar_salida" not in st.session_state:
    st.session_state["confirmar_salida"] = False
if "menu_seleccionado" not in st.session_state:
    st.session_state["menu_seleccionado"] = "Visualización"


# ==== Carga automática de base de datos ====
if st.session_state["df_memoria"] is None and os.path.exists(BASE_DATOS_PATH):
    try:
        hojas_disponibles = obtener_hojas(BASE_DATOS_PATH)
        if hojas_disponibles:
            hoja_excel = hojas_disponibles[0]
            df_base = pd.read_excel(BASE_DATOS_PATH, sheet_name=hoja_excel)
            if 'Call Open Interest' not in df_base.columns or 'Put Open Interest' not in df_base.columns:
                df_base.rename(columns={
                    'Open Interest': 'Call Open Interest',
                    'Open Interest.1': 'Put Open Interest'
                }, inplace=True)
            df_base['Expiration Date'] = pd.to_datetime(df_base['Expiration Date'])
            df_base['Fecha de Extracción'] = pd.to_datetime(df_base['Fecha de Extracción']).dt.date
            st.session_state["df_memoria"] = df_base
            st.session_state["nombre_base_activa"] = BASE_DATOS_PATH
            st.session_state["hojas_disponibles"] = hojas_disponibles
            st.session_state["nombre_hoja_excel"] = hoja_excel
    except Exception:
        st.session_state["df_memoria"] = None
        st.session_state["nombre_base_activa"] = None

# ===================== CARGAR CSS =====================
if os.path.exists("style.css"):
    with open("style.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ===================== MENÚ PRINCIPAL =====================
with st.sidebar:
    if st.session_state.get("hojas_disponibles"):
        st.markdown("<h4 style='color: white; margin-bottom: 0.5rem;'>📄 Selección de hoja activa</h4>",unsafe_allow_html=True)

        hoja_seleccionada = st.selectbox(
            "Selecciona hoja activa",  # Etiqueta accesible
            st.session_state["hojas_disponibles"],
            index=st.session_state["hojas_disponibles"].index(
                st.session_state.get("nombre_hoja_excel", "Datos")
            ),
            key="hoja_activa",
            label_visibility="collapsed"  # Oculta visualmente la etiqueta, pero mantiene accesibilidad
        )

    # Detectar cambio de hoja y actualizar df_memoria si es necesario
    if st.session_state.get("hojas_disponibles") and hoja_seleccionada != st.session_state.get("nombre_hoja_excel"):
        st.session_state["nombre_hoja_excel"] = hoja_seleccionada
        try:
            df = pd.read_excel(BASE_DATOS_PATH, sheet_name=hoja_seleccionada)
            st.session_state["df_memoria"] = df
            st.rerun()
        except Exception as e:
            st.error(f"No se pudo cargar la hoja seleccionada: {str(e)}")

    # Limpiar la marca de cambio pendiente
    if "cambio_hoja_pendiente" in st.session_state:
        del st.session_state["cambio_hoja_pendiente"]


 # Solo mostrar el menú si no estamos en modo de confirmación de salida
    if not st.session_state.get("confirmar_salida", False):
        # Obtener el índice actual basado en el menú guardado
        opciones_menu = ["Visualización", "Estadísticas", "Cargar Datos", "Configuración"]
        try:
            indice_actual = opciones_menu.index(st.session_state["menu_seleccionado"])
        except ValueError:
            indice_actual = 0
            st.session_state["menu_seleccionado"] = opciones_menu[0]
        
        selected = option_menu(
            "Menú Principal",
            opciones_menu,
            icons=["bar-chart", "graph-up", "upload", "gear"],
            menu_icon="cast",
            default_index=indice_actual,
        )
        
        # Actualizar el estado del menú seleccionado
        st.session_state["menu_seleccionado"] = selected
    else:
        # Mantener la selección anterior cuando estamos en confirmación
        selected = st.session_state.get("menu_seleccionado", "Visualización")

    # ===================== BOTÓN DE SALIR =====================
    st.markdown("---")
    
    # Mostrar confirmación de salida si está activada
    if st.session_state.get("confirmar_salida", False):
        st.markdown("#### ⚠️ ¿Seguro que desea salir?")
        col_si, col_no = st.columns(2)
        
        with col_si:
            if st.button("✅ Sí, salir", type="primary", use_container_width=True):
                st.balloons()
                st.success("¡Cerrando aplicación!")
                
                # Mostrar mensaje final
                st.markdown("""
                    <div style="text-align: center; margin-top: 20px; padding: 20px; 
                                background-color: #f0f2f6; border-radius: 10px;">
                        <h3>✅ Aplicación cerrada correctamente</h3>
                        <p><strong>Para cerrar completamente:</strong></p>
                        <p>1. Cierre esta pestaña del navegador</p>
                        <p>2. En la terminal, presione <kbd>Ctrl+C</kbd> para detener el servidor</p>
                    </div>
                """, unsafe_allow_html=True)
                
                # Detener completamente la ejecución
                os._exit(0)
        
        with col_no:
            if st.button("❌ Cancelar", use_container_width=True):
                st.session_state["confirmar_salida"] = False
                # No necesitamos st.rerun() aquí tampoco



    else:
        # Mostrar botón de salir normal
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.button("🚪 Salir", type="secondary", use_container_width=True):
                # Ya no necesitamos guardar el menú actual porque ya está en session_state
                st.session_state["confirmar_salida"] = True

# ===================== FUNCIONES =====================
def mostrar_mensaje(tipo, texto):
    if tipo == "success":
        st.toast(texto, icon="✅")
    elif tipo == "warning":
        st.toast(texto, icon="⚠️")
    elif tipo == "error":
        st.toast(texto, icon="❌")
    else:
        st.toast(texto, icon="ℹ️")

def cargar_xlsx(file):
    try:
        # Crear backup si existe una base de datos previa
        if os.path.exists(BASE_DATOS_PATH):
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"base_datos_backup_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copyfile(BASE_DATOS_PATH, backup_path)
            mostrar_mensaje("success", f"Copia de seguridad creada: {backup_path}")

        # Obtener las hojas del archivo cargado
        hojas = obtener_hojas(file)
        if not hojas:
            mostrar_mensaje("error", "No se encontraron hojas en el archivo.")
            return None

        hoja = hojas[0]
        df = pd.read_excel(file, sheet_name=hoja)

        # Renombrar columnas si es necesario
        if 'Call Open Interest' not in df.columns or 'Put Open Interest' not in df.columns:
            df.rename(columns={
                'Open Interest': 'Call Open Interest',
                'Open Interest.1': 'Put Open Interest'
            }, inplace=True)

        # Convertir fechas
        df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        df['Fecha de Extracción'] = pd.to_datetime(df['Fecha de Extracción']).dt.date

        # Guardar el archivo cargado como nueva base de datos
        with open(BASE_DATOS_PATH, "wb") as out_file:
            out_file.write(file.getbuffer())

        # Actualizar hojas disponibles (leer desde el nuevo archivo guardado)
        hojas_actualizadas = obtener_hojas(BASE_DATOS_PATH)

        # Actualizar estado de sesión
        st.session_state["df_memoria"] = df
        st.session_state["nombre_base_activa"] = file.name if hasattr(file, 'name') else BASE_DATOS_PATH
        st.session_state["nombre_hoja_excel"] = hoja
        st.session_state["hojas_disponibles"] = hojas_actualizadas

        mostrar_mensaje("success", f"Archivo Excel cargado correctamente y guardado como base de datos.")

        return df

    except Exception as e:
        mostrar_mensaje("error", f"Error al cargar y guardar el archivo Excel: {str(e)}")
        return None

def cargar_csv(file):
    try:
        df = pd.read_csv(file, skiprows=3, header=0, sep=',')
        df = df[['Expiration Date', 'Open Interest', 'Strike', 'Open Interest.1']]
        df.columns = ['Expiration Date', 'Call Open Interest', 'Strike', 'Put Open Interest']
        df.insert(0, 'Fecha de Extracción', pd.to_datetime('today').date())
        #df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        df['Expiration Date'] = pd.to_datetime(df['Expiration Date'])
        return df
    except Exception as e:
        mostrar_mensaje("error", f"Error al cargar el archivo CSV: {str(e)}")
        return None
    
def guardar_base_datos(df, hoja_destino, backup=False, backup_name=None):
    try:
        # Leer todas las hojas existentes, excepto la que será reemplazada
        hojas = {}
        if os.path.exists(BASE_DATOS_PATH):
            try:
                xls = pd.ExcelFile(BASE_DATOS_PATH, engine="openpyxl")
                for hoja in xls.sheet_names:
                    if hoja != hoja_destino:
                        hojas[hoja] = xls.parse(hoja)
            except Exception as e:
                mostrar_mensaje("error", f"No se pudieron leer las hojas existentes: {e}")
                hojas = {}

        # Guardar la hoja destino con sus nuevos datos
        hojas[hoja_destino] = df

        with pd.ExcelWriter(BASE_DATOS_PATH, engine="openpyxl", mode="w") as writer:
            for hoja, datos in hojas.items():
                datos.to_excel(writer, sheet_name=hoja, index=False)

        # Crear backup si se solicita
        if backup and os.path.exists(BASE_DATOS_PATH):
            backup_dir = "backups"
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = backup_name or f"base_datos_backup_{timestamp}.xlsx"
            backup_path = os.path.join(backup_dir, backup_filename)
            shutil.copyfile(BASE_DATOS_PATH, backup_path)
            mostrar_mensaje("success", f"Copia de seguridad completa creada: {backup_path}")

        # Leer nuevamente las hojas desde el archivo para evitar duplicados
        st.session_state["hojas_disponibles"] = obtener_hojas(BASE_DATOS_PATH)
        mostrar_mensaje("success", f"Datos guardados exitosamente en la hoja '{hoja_destino}'")

    except Exception as e:
        mostrar_mensaje("error", f"Error al guardar la base de datos: {str(e)}")

def generar_grafico_barras(df, fecha_vencimiento, fecha_carga=None):
    """Genera un gráfico de barras horizontales para los TOP 10 CALL y TOP 10 PUT con etiquetas de valor."""

    top_calls = df.nlargest(10, 'Call Open Interest')
    top_puts = df.nlargest(10, 'Put Open Interest')

    strikes_top = pd.concat([top_calls, top_puts]).drop_duplicates(subset='Strike').sort_values('Strike')

    max_call = strikes_top['Call Open Interest'].max()
    max_put = strikes_top['Put Open Interest'].max()
    max_oi = max(max_call, max_put)
    x_buffer = max_oi * 0.2

    fig, ax = plt.subplots(figsize=(12, 7))

    if fecha_carga:
        plt.suptitle(
            f"{hoja_seleccionada} - Open Interest (TOP 10 CALL & TOP 10 PUT)\n"
            f"Vencimiento: {fecha_vencimiento.strftime('%Y-%m-%d')} | "
            f"Extracción: {fecha_carga.strftime('%Y-%m-%d')}",
            fontsize=15
        )
    else:
        plt.suptitle(
            f"{hoja_seleccionada} - Open Interest (TOP 10 CALL & TOP 10 PUT)\n"
            f"Vencimiento: {fecha_vencimiento.strftime('%Y-%m-%d')}",
            fontsize=15
        )

    plt.title(f"Informe generado el: {datetime.now().strftime('%Y-%m-%d')}", fontsize=11, style='italic', loc='center')

    # Dibujar barras
    bar_labels = strikes_top['Strike'].astype(str)
    call_values = -strikes_top['Call Open Interest']
    put_values = strikes_top['Put Open Interest']

    bars_call = ax.barh(bar_labels, call_values, color='#4CAF50', alpha=0.8, label='CALL OI')
    bars_put = ax.barh(bar_labels, put_values, color='#F44336', alpha=0.8, label='PUT OI')

    ax.axvline(0, color='black', linestyle='-', linewidth=1.5)

    # Añadir valores CALL
    for bar in bars_call:
        width = bar.get_width()
        if width != 0:
            ax.annotate(
                f'{abs(int(width)):,}',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(-5, 0),
                textcoords="offset points",
                ha='right',
                va='center',
                color='#2E7D32',
                fontsize=10
            )

    # Añadir valores PUT
    for bar in bars_put:
        width = bar.get_width()
        if width != 0:
            ax.annotate(
                f'{int(width):,}',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(5, 0),
                textcoords="offset points",
                ha='left',
                va='center',
                color='#C62828',
                fontsize=10
            )

    ax.set_xlim(-max_call - x_buffer, max_put + x_buffer)
    ax.set_xlabel('Open Interest', labelpad=10)
    ax.set_ylabel('Strike Price', labelpad=10)
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.1), ncol=2, framealpha=1)
    ax.grid(axis='y', linestyle='--', alpha=0.4)

    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)

    fig.tight_layout()
    return fig


def generar_grafico_evolucion_strike(df, fecha_vencimiento, strike_seleccionado):
    """Genera un gráfico de líneas mostrando la evolución del Open Interest para un strike específico."""
    
    # Filtrar datos para el vencimiento y strike específicos
    df_filtrado = df[
        (df['Expiration Date'] == fecha_vencimiento) & 
        (df['Strike'] == strike_seleccionado)
    ].copy()
    
    if df_filtrado.empty:
        return None
    
    # Ordenar por fecha de extracción
    df_filtrado = df_filtrado.sort_values('Fecha de Extracción')
    
    # Crear el gráfico
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Configurar título
    plt.suptitle(
        f"{st.session_state['nombre_hoja_excel']} - Evolución Open Interest\n"
        f"Strike: {strike_seleccionado} | Vencimiento: {fecha_vencimiento.strftime('%Y-%m-%d')}",
        fontsize=15
    )
    
    plt.title(f"Informe generado el: {datetime.now().strftime('%Y-%m-%d')}", 
              fontsize=11, style='italic', loc='center')
    
    # Preparar datos para el gráfico
    fechas = df_filtrado['Fecha de Extracción']
    call_oi = df_filtrado['Call Open Interest']
    put_oi = df_filtrado['Put Open Interest']
    
    # Dibujar líneas
    ax.plot(fechas, call_oi, marker='o', linewidth=2.5, markersize=6, 
            color='#4CAF50', label='CALL Open Interest', alpha=0.8)
    ax.plot(fechas, put_oi, marker='s', linewidth=2.5, markersize=6, 
            color='#F44336', label='PUT Open Interest', alpha=0.8)
    
    # Agregar valores en los puntos
    for i, (fecha, call_val, put_val) in enumerate(zip(fechas, call_oi, put_oi)):
        ax.annotate(f'{int(call_val):,}', 
                   xy=(fecha, call_val), 
                   xytext=(0, 10), 
                   textcoords="offset points",
                   ha='center', va='bottom',
                   fontsize=9, color='#2E7D32',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.7))
        
        ax.annotate(f'{int(put_val):,}', 
                   xy=(fecha, put_val), 
                   xytext=(0, -15), 
                   textcoords="offset points",
                   ha='center', va='top',
                   fontsize=9, color='#C62828',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.7))
    
    # Configurar ejes
    ax.set_xlabel('Fecha de Extracción', labelpad=10)
    ax.set_ylabel('Open Interest', labelpad=10)
    ax.legend(loc='upper left', framealpha=1)
    ax.grid(True, linestyle='--', alpha=0.4)
    
    # Rotar etiquetas del eje x para mejor legibilidad
    plt.xticks(rotation=45)
    
    # Quitar bordes superiores y derechos
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    
    fig.tight_layout()
    return fig

# ===================== FUNCIONALIDAD =====================
# Solo mostrar contenido si no estamos en modo de confirmación de salida
if not st.session_state.get("confirmar_salida", False):
    if selected == "Cargar Datos":
        st.markdown("<h2 class='fade-in'>Cargar Datos</h2>", unsafe_allow_html=True)
        opcion_menu = st.radio("Seleccione una opción:", ("Cargar nueva base de datos (Excel)", "Ampliar base de datos existente (CSV)"))

        if opcion_menu == "Cargar nueva base de datos (Excel)":
            uploaded_file = st.file_uploader("Seleccione archivo de datos (XLSX)", type=["xlsx"])

            if uploaded_file:
                df = cargar_xlsx(uploaded_file)
                if df is None or df.empty:
                    mostrar_mensaje("error", "No se pudieron cargar los datos. Verifique el formato del archivo.")
                    st.stop()

                hoja_cargada = st.session_state.get("nombre_hoja_excel", "Datos")

                # Guardar la hoja cargada
                #guardar_base_datos(df, hoja_destino=hoja_cargada)

                # Actualizar el estado después de guardar
                st.session_state["df_memoria"] = df
                st.session_state["nombre_base_activa"] = uploaded_file.name

            
        else:
            if st.session_state["df_memoria"] is None:
                mostrar_mensaje("error", "Primero debe cargar una base de datos Excel para poder ampliarla con CSV.")
                st.stop()

            if "fecha_extraccion_csv" not in st.session_state:
                st.session_state["fecha_extraccion_csv"] = datetime.today().date()

            st.session_state["fecha_extraccion_csv"] = st.date_input(
                "Seleccione la fecha de extracción para los nuevos datos:",
                value=st.session_state["fecha_extraccion_csv"]
            )

            uploaded_csv = st.file_uploader("Seleccione archivo de datos (CSV)", type=["csv"])

            if uploaded_csv is not None:
                hoja_csv = st.selectbox(
                    "Selecciona la hoja donde agregar los datos:",
                    st.session_state.get("hojas_disponibles", ["Datos"])
                )

                nuevos_datos = cargar_csv(uploaded_csv)
                if nuevos_datos is None or nuevos_datos.empty:
                    mostrar_mensaje("error", "No se pudieron cargar los datos CSV.")
                    st.stop()

                fecha_elegida = st.session_state["fecha_extraccion_csv"]
                nuevos_datos['Fecha de Extracción'] = fecha_elegida

                if st.button("📥 Confirmar carga en la base de datos"):
                    try:
                        base_actual = pd.read_excel(BASE_DATOS_PATH, sheet_name=hoja_csv)
                        base_actual['Expiration Date'] = pd.to_datetime(base_actual['Expiration Date'])
                        base_actual['Fecha de Extracción'] = pd.to_datetime(base_actual['Fecha de Extracción']).dt.date

                        combinaciones_existentes = base_actual[['Fecha de Extracción', 'Expiration Date']].drop_duplicates()
                        conflictivas = nuevos_datos.merge(combinaciones_existentes, on=['Fecha de Extracción', 'Expiration Date'], how='inner')

                        if not conflictivas.empty:
                            mostrar_mensaje("error", f"Ya existen datos para esa combinación en la hoja '{hoja_csv}'.")
                            st.stop()

                        # Crear copia de seguridad antes de modificar
                        backup_name = f"base_datos_backup_{hoja_csv}_{fecha_elegida.strftime('%Y%m%d')}.xlsx"
                        guardar_base_datos(base_actual, hoja_destino=hoja_csv, backup=True, backup_name=backup_name)

                        base_merged = pd.concat([base_actual, nuevos_datos]).drop_duplicates(
                            ['Fecha de Extracción', 'Expiration Date', 'Strike'],
                            keep='last'
                        )

                        # Guardar solo la hoja seleccionada
                        guardar_base_datos(base_merged, hoja_destino=hoja_csv)

                        # Actualizar solo si se guardó correctamente
                        st.session_state["df_memoria"] = base_merged
                        st.session_state["nombre_base_activa"] = uploaded_csv.name
                        st.session_state["nombre_hoja_excel"] = hoja_csv

                        mostrar_mensaje("success", f"Datos agregados exitosamente a la hoja '{hoja_csv}'.")

                    except Exception as e:
                        mostrar_mensaje("error", f"Error al actualizar la hoja: {str(e)}")


    elif selected == "Visualización":
        st.markdown("<h2 class='fade-in'>Visualización</h2>", unsafe_allow_html=True)
        df = st.session_state["df_memoria"]

        if df is not None and not df.empty:
            # Filtrar fechas de extracción que tengan al menos una fecha de vencimiento asociada
            fechas_extraccion_validas = df.groupby('Fecha de Extracción')['Expiration Date'].nunique()
            fechas_extraccion = sorted(fechas_extraccion_validas[fechas_extraccion_validas > 0].index, reverse=True)

            # Inicialmente se puede dejar las fechas de vencimiento vacías (se actualizarán luego)
            fechas_vencimiento = []


            col1, col2 = st.columns(2)

            with col1:
                fecha_extraccion = st.selectbox(
                    "Seleccione fecha de extracción:",
                    fechas_extraccion,
                    format_func=lambda x: x.strftime('%Y-%m-%d')
                )

            with col2:
                fechas_vencimiento = sorted(df[df['Fecha de Extracción'] == fecha_extraccion]['Expiration Date'].unique())
                fecha_vencimiento = st.selectbox(
                    "Seleccione fecha de vencimiento:",
                    fechas_vencimiento,
                    format_func=lambda x: x.strftime('%Y-%m-%d')
                )

            df_filtrado = df[(df['Fecha de Extracción'] == fecha_extraccion) & (df['Expiration Date'] == fecha_vencimiento)]

            if not df_filtrado.empty:
                st.subheader(f"{st.session_state['nombre_hoja_excel']} - Open Interest\nExtracción: {fecha_extraccion} | Vencimiento: {fecha_vencimiento}")
                fig = generar_grafico_barras(df_filtrado, fecha_vencimiento, fecha_extraccion)
                st.pyplot(fig)

                buf = io.BytesIO()
                fig.savefig(buf, format="png", bbox_inches='tight')
                buf.seek(0)

                file_name = (
                    f"{st.session_state['nombre_hoja_excel']} - IO vencimiento ({fecha_vencimiento.strftime('%Y-%m-%d')}) "
                    f"- extraccion ({fecha_extraccion.strftime('%Y-%m-%d')}).png"
                )
                
                st.download_button("📥 Descargar imagen", data=buf, file_name=file_name, mime="image/png")

                if st.checkbox("Mostrar tabla de datos completos", value=False):
                    st.dataframe(df_filtrado[['Strike', 'Call Open Interest', 'Put Open Interest']].sort_values('Strike'))
            else:
                mostrar_mensaje("warning", "No hay datos disponibles para los filtros seleccionados.")
        else:
            st.info("No se ha cargado ninguna base de datos. Por favor, use la opción 'Cargar Datos'.")

    # ESTADISTICAS
    
    elif selected == "Estadísticas":
        st.markdown("<h2 class='fade-in'>Estadísticas</h2>", unsafe_allow_html=True)
        df = st.session_state["df_memoria"]

        if df is not None and not df.empty:
            # Crear dos columnas: controles a la izquierda, gráfico a la derecha
            col_controles, col_grafico = st.columns([1, 2])
            
            with col_controles:
                st.markdown("### Selección de parámetros")
                
                # Obtener fechas de vencimiento únicas (ordenadas de mayor a menor)
                fechas_vencimiento_unicas = sorted(df['Expiration Date'].unique(), reverse=True)
                
                fecha_vencimiento_stats = st.selectbox(
                    "Seleccione fecha de vencimiento:",
                    fechas_vencimiento_unicas,
                    format_func=lambda x: x.strftime('%Y-%m-%d'),
                    key="fecha_vencimiento_stats"
                )
                
                # Obtener strikes disponibles para la fecha de vencimiento seleccionada
                strikes_disponibles = sorted(
                    df[df['Expiration Date'] == fecha_vencimiento_stats]['Strike'].unique()
                )
                
                strike_seleccionado = st.selectbox(
                    "Seleccione strike:",
                    strikes_disponibles,
                    key="strike_seleccionado_stats"
                )

            with col_grafico:
                st.markdown("### Evolución del Open Interest")
                
                # Generar y mostrar el gráfico
                fig_evolucion = generar_grafico_evolucion_strike(
                    df, fecha_vencimiento_stats, strike_seleccionado
                )
                
                if fig_evolucion is not None:
                    st.pyplot(fig_evolucion)
                    
                    # Botón de descarga
                    buf = io.BytesIO()
                    fig_evolucion.savefig(buf, format="png", bbox_inches='tight')
                    buf.seek(0)
                    
                    file_name = (
                        f"{st.session_state['nombre_hoja_excel']} - Evolucion Strike {strike_seleccionado} "
                        f"- Vencimiento ({fecha_vencimiento_stats.strftime('%Y-%m-%d')}).png"
                    )
                    
                    st.download_button(
                        "📥 Descargar gráfico de evolución", 
                        data=buf, 
                        file_name=file_name, 
                        mime="image/png"
                    )
                    
                    # Mostrar tabla de datos si se solicita
                    if st.checkbox("Mostrar datos históricos", value=False, key="mostrar_datos_stats"):

                        df_strike = df[
                            (df['Expiration Date'] == fecha_vencimiento_stats) & 
                            (df['Strike'] == strike_seleccionado)
                        ].copy()

                        df_mostrar = df_strike[['Fecha de Extracción', 'Call Open Interest', 'Put Open Interest']].copy()
                        df_mostrar = df_mostrar.sort_values('Fecha de Extracción', ascending=False)
                        df_mostrar['Fecha de Extracción'] = df_mostrar['Fecha de Extracción'].apply(
                            lambda x: x.strftime('%Y-%m-%d')
                        )
                        st.dataframe(df_mostrar, use_container_width=True)
                else:
                    st.warning("No hay datos disponibles para el strike y fecha de vencimiento seleccionados.")
        else:
            st.info("No se ha cargado ninguna base de datos. Por favor, use la opción 'Cargar Datos'.")

    # CONFIGURACION
    elif selected == "Configuración":
        st.markdown("<h2 class='fade-in'>Configuración</h2>", unsafe_allow_html=True)
        st.write("Opciones de configuración próximamente.")

    

   
